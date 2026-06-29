# nextsoft_export_to_drive.py
# Fluxo:
# 1) Login (resiliente: retries, timeouts maiores, user-agent real)
# 2) Troca de loja (via secret APPNEXT_LOJA_DESTINO)
# 3) Vendas > Vendedor Analítico
# 4) Abre filtros (estabiliza grid)
# 5) Tenta exportar; se URL direta não for capturada (blob/POST), FALLBACK:
#    refaz a chamada de dados em janelas mensais e gera o Excel localmente
# 6) Envia o arquivo ao Google Drive via rclone

import os
import sys
import subprocess
import traceback
import shutil
import json
import re
import threading           # <<< NOVO: usado pelo watchdog global
import time                # <<< NOVO: usado pelo watchdog global
from pathlib import Path
from datetime import datetime, timedelta, timezone
from urllib.parse import urlparse, parse_qs, urlencode

from dotenv import load_dotenv
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# ================= LOG =================
def log(msg: str):
    now = datetime.now().strftime("%H:%M:%S")
    print(f"[{now}] {msg}", flush=True)


def step(title: str):
    log("=" * 70)
    log(title)
    log("=" * 70)


# ============== WATCHDOG GLOBAL (NOVO) ==============
# Mata o processo se ele exceder o tempo máximo, não importa onde
# o Playwright esteja travado. Roda numa thread separada (daemon).
def start_watchdog(max_seconds: int):
    def _killer():
        time.sleep(max_seconds)
        log(f"### TIMEOUT GLOBAL: excedeu {max_seconds}s ({max_seconds // 60} min). Abortando. ###")
        # 124 é o código convencional de timeout; faz o job FALHAR no Actions
        os._exit(124)
    threading.Thread(target=_killer, daemon=True).start()
    log(f"[watchdog] ativo: aborta em {max_seconds}s ({max_seconds // 60} min).")


load_dotenv()

# ================ CONFIG ================
REDE = os.getenv("APPNEXT_REDE", "").strip()
USER = os.getenv("APPNEXT_USER", "").strip()
PASS = os.getenv("APPNEXT_PASS", "").strip()

DRIVE_REMOTE = os.getenv("DRIVE_REMOTE", "GDRIVE:")
DRIVE_FOLDER_ID = os.getenv("DRIVE_FOLDER_ID", "").strip()
DRIVE_FILE_NAME = os.getenv("DRIVE_FILE_NAME", "importacaoA.xlsx").strip()

# Tempo máximo total do processo, em segundos (default 7 min = 420s).
# Pode ser sobrescrito via secret/variável MAX_RUNTIME_SECONDS.
MAX_RUNTIME_SECONDS = int(os.getenv("MAX_RUNTIME_SECONDS", "420"))

# ---- Datas ----
# Lógica:
#  - Se DATA_INICIO for definida explicitamente no ambiente, usa ela.
#  - Senão, se DIAS_JANELA for definida, calcula uma janela deslizante
#    (ex.: últimos 60 dias) -> mantém o tempo de execução estável.
#  - Senão, mantém o comportamento antigo (data fixa 01/06/2025).
DIAS_JANELA = os.getenv("DIAS_JANELA", "").strip()
_data_inicio_env = os.getenv("DATA_INICIO", "").strip()

if _data_inicio_env:
    DATA_INICIO = _data_inicio_env
elif DIAS_JANELA:
    _dias = int(DIAS_JANELA)
    DATA_INICIO = (datetime.now() - timedelta(days=_dias)).strftime("%d/%m/%Y")
else:
    DATA_INICIO = "01/06/2025"  # comportamento original preservado

DATA_FIM = os.getenv("DATA_FIM", "").strip()  # vazio => hoje 23:59:59

TEMPLATE_XLSX = os.getenv("TEMPLATE_XLSX", "importacaoA.xlsx")
TEMPLATE_HEADER_ROW = int(os.getenv("TEMPLATE_HEADER_ROW", "2"))

# Forçar lojaId diretamente no fallback (se quiser)
LOJA_ID_DESTINO = os.getenv("APPNEXT_LOJA_ID_DESTINO", "").strip()

RCLONE_PATH = os.getenv("RCLONE_PATH") or shutil.which("rclone") or r"C:\rclone\rclone.exe"
if not Path(RCLONE_PATH).exists() and shutil.which("rclone") is None:
    log("ERRO: rclone não encontrado. Ajuste RCLONE_PATH ou adicione ao PATH.")
    sys.exit(1)

if not (REDE and USER and PASS):
    log("ERRO: Preencha APPNEXT_REDE / APPNEXT_USER / APPNEXT_PASS")
    sys.exit(1)

TS = datetime.now().strftime("%Y%m%d_%H%M%S")
LOCAL_OUT = Path.cwd() / f"export_nextsoft_{TS}.xlsx"
LOGIN_URL = "https://www.appnext.com.br/#/login"

# -------- datas ----------
def to_dt(s: str, end=False):
    try:
        d = datetime.strptime(s, "%d/%m/%Y")
        return d.replace(hour=23, minute=59, second=59) if end else d
    except Exception:
        return None


if not DATA_FIM:
    DATA_FIM = datetime.now().strftime("%d/%m/%Y")

INI_DT = to_dt(DATA_INICIO) or datetime(2025, 6, 1)
FIM_DT = to_dt(DATA_FIM, end=True) or datetime.now().replace(hour=23, minute=59, second=59)

# ISO com Z (mesmo padrão visto no DevTools)
FMT_JSON_INI = INI_DT.strftime("%Y-%m-%dT%H:%M:%S.000Z")
FMT_JSON_FIM = FIM_DT.strftime("%Y-%m-%dT%H:%M:%S.000Z")
FMT_BR_INI = INI_DT.strftime("%d/%m/%Y, %H:%M:%S")
FMT_BR_FIM = FIM_DT.strftime("%d/%m/%Y, %H:%M:%S")

log(f"[datas] período: {FMT_BR_INI}  ->  {FMT_BR_FIM}")

# -------- seletores ----------
SEL = {
    # login
    "rede": "input[placeholder='Rede']",
    "email": "input[placeholder='Email']",
    "senha": "input[placeholder='Senha']",
    "entrar": "button:has-text('Entrar'), button:has-text('Login')",
    # menu
    "menu_vendas": "nav >> text=Vendas, header >> text=Vendas, a[href*='vendas']:has-text('Vendas'), button:has-text('Vendas')",
    "item_vendedor_analitico": "a[href*='vendedor-analitico'], a:has-text('Vendedor Anal'), [role='menuitem']:has-text('Vendedor Anal'), li:has-text('Vendedor Anal')",
    # tela alvo
    "titulo_rel": "text=Listagem de Vendedor Anal",
    # filtros / grid / export
    "btn_filtros": "[title*='Filtro'], [aria-label*='Filtro'], button:has(i.mdi-filter), button:has(i.fa-filter), button:has-text('Filtros')",
    "pane_filtros": "#filtrosForm",
    "btn_atualizar": "button:has-text('Atualizar Filtros')",
    "grid_row": "table tbody tr",
    "excel_btn": "#dataTableButtons button.buttons-excel, button.buttons-excel",
}

# =============== HELPERS ===============
def rclone_copy_latest(local_file: Path):
    cmd = [
        RCLONE_PATH,
        "copyto",
        str(local_file),
        f"{DRIVE_REMOTE}{DRIVE_FILE_NAME}",
        "--drive-root-folder-id",
        DRIVE_FOLDER_ID,
        "-v",
    ]
    log(f"rclone -> {' '.join(cmd)}")
    res = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8", errors="replace")
    if res.stdout.strip():
        print(res.stdout)
    if res.returncode != 0:
        print(res.stderr, file=sys.stderr)
        raise RuntimeError("Falha ao atualizar o arquivo no Drive.")


def _click_first(page, selectors):
    for sel in selectors:
        try:
            loc = page.locator(sel).first
            if loc.count():
                loc.scroll_into_view_if_needed(timeout=1500)
                loc.click(timeout=4000, force=True)
                return True
        except Exception:
            pass
    return False


def wait_for_titulo_rel(page, timeout=40000):
    """Aguarda qualquer variação do título do relatório Vendedor Analítico."""
    sels = [
        SEL["titulo_rel"],
        "h1:has-text('Vendedor')",
        "h2:has-text('Vendedor')",
        ".page-title:has-text('Vendedor')",
        "text=Vendedor Anal",
    ]
    deadline = timeout
    for sel in sels:
        try:
            page.wait_for_selector(sel, timeout=deadline)
            return
        except Exception:
            pass
    raise TimeoutError(f"Título de Vendedor Analítico não encontrado após {timeout}ms")


def goto_vendedor_analitico_via_menu(page):
    # --- tentativa 1: navegação direta via URL ---
    base = page.url.split("/#/")[0] if "/#/" in page.url else "https://www.appnext.com.br"
    direct_url = f"{base}/#/loja/vendas/vendedor-analitico"
    try:
        log(f"[nav] tentando URL direta -> {direct_url}")
        page.goto(direct_url, wait_until="load", timeout=60_000)
        try:
            page.wait_for_url("**/loja/vendas/**vendedor**", timeout=15000)
        except PWTimeout:
            pass
        try:
            wait_for_titulo_rel(page, timeout=15000)
            return  # sucesso via URL direta
        except Exception:
            pass
    except Exception as e:
        log(f"[nav] URL direta falhou: {e}")

    # --- tentativa 2: via menu ---
    for sel in [s.strip() for s in SEL["menu_vendas"].split(",")]:
        try:
            loc = page.locator(sel).first
            if loc.count():
                try:
                    loc.hover()
                except Exception:
                    pass
                loc.click(timeout=4000, force=True)
                break
        except Exception:
            continue
    for sel in [s.strip() for s in SEL["item_vendedor_analitico"].split(",")]:
        try:
            it = page.locator(sel).first
            if it.count():
                it.click(timeout=6000, force=True)
                break
        except Exception:
            continue
    # confirma rota/título
    try:
        page.wait_for_url("**/loja/vendas/**vendedor**", timeout=30000)
    except PWTimeout:
        wait_for_titulo_rel(page, timeout=20000)


def open_filters_pane(page):
    # se já estiver aberto, retorna
    try:
        page.wait_for_selector(f"{SEL['pane_filtros']}.show", timeout=1500)
        return
    except Exception:
        pass
    # clica no botão/ícone de filtros
    for sel in [s.strip() for s in SEL["btn_filtros"].split(",")]:
        try:
            loc = page.locator(sel).first
            if loc.count():
                loc.scroll_into_view_if_needed(timeout=1500)
                loc.click(timeout=2000, force=True)
                break
        except Exception:
            continue
    # espera abrir
    try:
        page.wait_for_selector(f"{SEL['pane_filtros']}.show", timeout=4000)
    except Exception:
        # força abrir (caso offcanvas)
        page.evaluate(
            """idSel => { const p=document.querySelector(idSel); if(p){p.classList.add('show'); p.style.display='block';} }""",
            SEL["pane_filtros"],
        )

# --------- CAPTURA DO FILTRO (debug/auxiliar) ----------
class Captured:
    filtro_url = None
    filtro_method = None
    filtro_ct = None
    filtro_body = None
    filtro_headers = None


def _clean_headers(h: dict) -> dict:
    bad = {"content-length", "host", "origin", "referer"}
    return {k: v for k, v in (h or {}).items() if k.lower() not in bad}


def hook_filter_capture(page):
    def on_request(req):
        url = (req.url or "").lower()
        if Captured.filtro_url is None and req.resource_type in ("xhr", "fetch"):
            # request principal do relatório (GET com query)
            is_report_data = ("vendedor" in url and "analit" in url) or "relatoriovendedoranalitico" in url
            if is_report_data:
                hdrs = _clean_headers(req.headers or {})
                body = ""
                try:
                    body = req.post_data or ""
                except Exception:
                    pass
                Captured.filtro_url = req.url
                Captured.filtro_method = (req.method or "GET").upper()
                Captured.filtro_body = body
                Captured.filtro_ct = hdrs.get("content-type", "")
                Captured.filtro_headers = hdrs
                log(f"[capture] filtro de DADOS {Captured.filtro_method} -> {req.url}")

    page.on("request", on_request)

# --------- CAPTURA DO EXPORT (se existir URL direta) ----------
class CapturedExport:
    url = None
    method = None
    headers = None


def hook_export_capture(page):
    def on_request(req):
        url = (req.url or "").lower()
        if req.resource_type in ("xhr", "fetch", "document"):
            if ("excel" in url or "export" in url or "xlsx" in url) and ("vendedor" in url or "analit" in url):
                CapturedExport.url = req.url
                CapturedExport.method = (req.method or "GET").upper()
                CapturedExport.headers = _clean_headers(req.headers or {})
                log(f"[capture] EXPORT {CapturedExport.method} -> {req.url}")

    page.on("request", on_request)

# --------- REPLAY DO EXPORT (se capturado) ---------
def replay_export_with_dates(page, destino_path: Path):
    if not CapturedExport.url:
        raise RuntimeError("URL de exportação não foi capturada. Clique 1x no botão Excel para eu aprender a URL.")
    parsed = urlparse(CapturedExport.url)
    qs = parse_qs(parsed.query)
    qs["dataInicial"] = [FMT_JSON_INI]
    qs["dataFinal"] = [FMT_JSON_FIM]
    target_url = f"{parsed.scheme}://{parsed.netloc}{parsed.path}?{urlencode(qs, doseq=True)}"
    log(f"[export-replay] GET -> {target_url}")
    # timeout reduzido de 180s -> 90s
    with page.expect_download(timeout=90_000) as dl:
        page.evaluate(
            """
            (url) => { const a=document.createElement('a'); a.href=url; a.target='_blank';
                       document.body.appendChild(a); a.click(); setTimeout(()=>a.remove(),500); }
            """,
            target_url,
        )
    download = dl.value
    download.save_as(destino_path.as_posix())
    log(f"Arquivo exportado (período aplicado): {destino_path.name}")

# --------- FALLBACK: baixar JSON em janelas mensais ----------
def _janelas_mensais(ini_dt, fim_dt):
    janelas, cur = [], ini_dt
    while cur <= fim_dt:
        if cur.month == 12:
            prox = cur.replace(year=cur.year + 1, month=1, day=1)
        else:
            prox = cur.replace(month=cur.month + 1, day=1)
        fim_janela = min(prox - timedelta(seconds=1), fim_dt)
        janelas.append((cur, fim_janela))
        cur = prox
    return janelas


def fetch_report_json_with_dates(page):
    if not Captured.filtro_url:
        raise RuntimeError("Endpoint de dados não capturado ainda.")
    parsed = urlparse(Captured.filtro_url)
    base_qs = parse_qs(parsed.query)
    if LOJA_ID_DESTINO:
        base_qs["lojaId"] = [LOJA_ID_DESTINO]
    headers = {**_clean_headers(Captured.filtro_headers or {})}

    js_code = """
    async ([url, headers]) => {
        const res = await fetch(url, { method: "GET", headers });
        const text = await res.text();
        return { ok: res.ok, status: res.status, text };
    }
    """

    todos = []
    for jini, jfim in _janelas_mensais(INI_DT, FIM_DT):
        qs = dict(base_qs)
        qs["dataInicial"] = [jini.strftime("%Y-%m-%dT%H:%M:%S.000Z")]
        qs["dataFinal"]   = [jfim.strftime("%Y-%m-%dT%H:%M:%S.000Z")]
        target_url = f"{parsed.scheme}://{parsed.netloc}{parsed.path}?{urlencode(qs, doseq=True)}"
        log(f"[dados-replay] {jini:%d/%m/%Y} -> {jfim:%d/%m/%Y}")

        result = None
        # retries reduzidos: 2 tentativas e backoff menor (cabe no orçamento de tempo)
        for tentativa in range(1, 3):
            result = page.evaluate(js_code, [target_url, headers])
            if result.get("ok"):
                break
            log(f"  HTTP {result.get('status')} (tentativa {tentativa}/2)")
            page.wait_for_timeout(1000 * tentativa)

        if not result or not result.get("ok"):
            raise RuntimeError(
                f"Falha na janela {jini:%d/%m} - {jfim:%d/%m}: "
                f"HTTP {result.get('status') if result else '??'}"
            )

        try:
            parcial = json.loads(result["text"])
        except Exception as e:
            raise RuntimeError(f"JSON inválido na janela {jini:%d/%m}: {e}") from e

        if isinstance(parcial, list):
            todos.extend(parcial)
        elif isinstance(parcial, dict):
            for key in ("items", "data", "resultado", "rows", "content"):
                if isinstance(parcial.get(key), list):
                    todos.extend(parcial[key])
                    break
            else:
                todos.append(parcial)

    log(f"[dados-replay] total de linhas agregadas: {len(todos)}")
    return todos

# ----------------- GERAÇÃO DO EXCEL (idêntico ao importacaoA.xlsx) -----------------
# Larguras EXATAS extraídas do modelo
COL_WIDTHS = {
    "A": 32.4, "B": 18.9, "C": 13.5, "D": 16.2, "E": 25.65, "F": 18.9,
    "G": 54.0, "H": 40.5, "I": 52.65, "J": 54.0, "K": 13.5, "L": 18.9,
    "M": 22.95, "N": 14.85, "O": 16.2,
}

# Layout do arquivo final, na ordem das colunas A..O:
#   (cabeçalho, [campos candidatos vindos da API], tipo, env_fallback)
# tipo: "text" | "num" | "data_br"
# >>> AJUSTE os nomes de campo conforme o log "[debug] colunas do JSON normalizado:".
COLUNAS = [
    ("Loja",                ["loja"],                         "text",    "APPNEXT_LOJA_DESTINO"),
    ("CNPJ",                ["cnpj"],                         "text",    "APPNEXT_LOJA_CNPJ"),
    ("Data",                ["data"],                         "data_br", None),
    ("Número Cupom",        ["numeroCupom"],                  "text",    None),
    # CONFIRMADO pela amostra: valorTotal = valorUnitario * quantidadeItem,
    # logo 'quantidade' = total de itens do cupom (coluna E) e
    # 'quantidadeItem' = qtd da linha (coluna "Quantidade", K).
    ("Qtd. Itens no Cupom", ["quantidade", "quantidadeItem"], "num",     None),
    ("Código Produto",      ["codigo"],                       "text",    None),
    ("Descrição",           ["produto"],                      "text",    None),
    ("Grupo",               ["grupo"],                        "text",    None),
    ("Sub-Grupo",           ["subGrupo"],                     "text",    None),
    ("Categorias",          ["categorias"],                   "text",    None),
    ("Quantidade",          ["quantidadeItem", "quantidade"], "num",     None),
    ("Valor Unitário",      ["valorUnitario"],                "num",     None),
    ("Desconto Unitário",   ["desconto"],                     "num",     None),
    ("Valor Total",         ["valorTotal"],                   "num",     None),
    ("Vendedor",            ["vendedor"],                     "text",    None),
]


def _pick(rec, candidatos):
    """Primeiro campo presente e não vazio (trata NaN do pandas)."""
    for c in candidatos:
        if c in rec:
            v = rec[c]
            if v is None:
                continue
            if isinstance(v, float) and v != v:  # NaN
                continue
            if v == "":
                continue
            return v
    return None


def _num(v):
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return int(v) if float(v).is_integer() else v
    try:
        n = float(str(v).replace(",", "."))
        return int(n) if n.is_integer() else n
    except Exception:
        return v


def _data_br(v):
    """Converte a data da API para a string 'dd/mm/yyyy' (igual ao modelo).
    Se vier ISO com fuso (…Z), converte para horário de Brasília (-03:00)
    antes de extrair o dia, evitando virar o dia por causa do UTC."""
    if v is None or v == "":
        return ""
    s = str(v).strip()
    try:
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        if dt.tzinfo is not None:
            dt = dt.astimezone(timezone(timedelta(hours=-3)))
        return dt.strftime("%d/%m/%Y")
    except Exception:
        pass
    m = re.match(r"(\d{2})/(\d{2})/(\d{4})", s)
    if m:
        return s[:10]
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return f"{m.group(3)}/{m.group(2)}/{m.group(1)}"
    return s


def _texto(raw):
    if raw is None:
        return ""
    if isinstance(raw, list):
        partes = []
        for x in raw:
            if isinstance(x, dict):
                partes.append(str(x.get("nome") or x.get("descricao") or x))
            else:
                partes.append(str(x))
        return ", ".join(partes)
    return str(raw).strip()


def write_excel_from_json(data, out_path: Path):
    import pandas as pd

    # 1) extrai a lista de linhas
    if isinstance(data, list):
        rows = data
    elif isinstance(data, dict):
        rows = None
        for key in ("items", "data", "resultado", "rows", "content"):
            if isinstance(data.get(key), list):
                rows = data[key]
                break
        rows = rows if rows is not None else [data]
    else:
        rows = []

    df = pd.json_normalize(rows)
    log("[debug] colunas do JSON normalizado: " + ", ".join(map(str, df.columns)))
    if len(df):
        amostra = {k: df.iloc[0][k] for k in df.columns}
        log("[debug] 1a linha (amostra): "
            + json.dumps(amostra, ensure_ascii=False, default=str)[:1500])
    registros = df.to_dict("records")

    # 2) monta a planilha do zero, igual ao importacaoA.xlsx
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ts = datetime.now().strftime("%d/%m/%Y, %H:%M:%S")
    ws["A1"] = f"Vendas Por Vendedor - Listagem Vendedor Analítico ( {ts} )"
    ws["A1"].font = Font(name="Calibri", size=11)
    ws.merge_cells("A1:O1")

    hdr_font = Font(name="Calibri", size=11, bold=True)
    hdr_align = Alignment(horizontal="center")
    for j, (header, _c, _t, _f) in enumerate(COLUNAS, start=1):
        c = ws.cell(row=2, column=j, value=header)
        c.font = hdr_font
        c.alignment = hdr_align

    body_font = Font(name="Calibri", size=11)
    preenchidas = set()
    r = 3
    for rec in registros:
        for j, (header, cands, tipo, fb) in enumerate(COLUNAS, start=1):
            raw = _pick(rec, cands)
            if raw is None and fb:
                raw = os.getenv(fb, "") or None
            if raw is not None:
                preenchidas.add(header)

            if tipo == "data_br":
                val = _data_br(raw)
            elif tipo == "num":
                val = _num(raw)
            else:
                val = _texto(raw)

            cell = ws.cell(row=r, column=j, value=val)
            cell.font = body_font
        r += 1

    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    vazias = [h for (h, _c, _t, _f) in COLUNAS if h not in preenchidas]
    if vazias:
        log("[mapeamento] colunas SEM origem na API (saíram VAZIAS em TODAS as linhas, ajustar COLUNAS): "
            + ", ".join(vazias))

    wb.save(out_path.as_posix())
    log(f"Arquivo salvo no layout do modelo: {out_path.name} ({len(registros)} linhas de dados)")

# -------------- TROCAR LOJA --------------
def trocar_loja(page, loja_nome=None):
    """
    Tenta trocar a loja pelo menu superior.
    Estratégia: procurar qualquer botão/área com 'TEA SHOP' ou nome da cidade e clicar.
    Depois selecionar o item do menu pelo texto (aceita ponto final).
    Retorna True/False indicando sucesso.
    """
    alvo = (
        loja_nome
        or os.getenv("APPNEXT_LOJA_DESTINO")
        or "GOIANIA - TEA SHOP FLAMBOYANT"
    )
    alvo_regex = re.compile(rf"^{re.escape(alvo)}\.?$", re.IGNORECASE)

    step(f"2) Trocando loja → {alvo}")
    try:
        opened = _click_first(
            page,
            [
                "button:has-text('TEA SHOP')",
                "button:has-text('VILA MADALENA')",
                "button:has-text('FLAMBOYANT')",
                "text=/S[ÂA]O PAULO - TEA SHOP|GOI[ÂA]NIA - TEA SHOP/i",
                "[data-bs-toggle='dropdown']",
                ".dropdown-toggle",
                "xpath=(//*[contains(translate(normalize-space(.),'áãéíóúâêô','aaeiouaeo'),'TEA SHOP')])[1]",
            ],
        )

        if not opened:
            # último recurso: clicar em algo com ícone de loja (emoji/mdi) + caret
            opened = _click_first(
                page,
                [
                    "xpath=(//i[contains(@class,'store') or contains(@class,'shop')]/ancestor::*[self::button or self::a])[1]"
                ],
            )

        # Seleciona a opção pelo texto
        try:
            page.get_by_role("menuitem", name=alvo_regex).first.click(timeout=6000, force=True)
        except Exception:
            clicked = _click_first(
                page,
                [
                    f"text=^{alvo}$",
                    f"text=^{re.escape(alvo)}",
                    f"xpath=//*[normalize-space()='{alvo}'] | //*[starts-with(normalize-space(), '{alvo}')]",
                ],
            )
            if not clicked:
                raise RuntimeError("Não encontrei a opção da loja no dropdown.")

        log(f"Loja selecionada: {alvo}")
        try:
            page.wait_for_load_state("networkidle", timeout=12000)
        except Exception:
            pass

        return True

    except Exception as e:
        log(f"Falha ao trocar loja: {e}")
        return False

# -------- Navegação resiliente: goto com retries --------
def goto_login_with_retries(page, url, tries=3):
    # tries reduzido de 5 -> 3 e timeout de 240s -> 60s por tentativa.
    # Antes, no pior caso, o login sozinho podia consumir ~20 min.
    alt_urls = {url, url.replace("https://www.", "https://")}
    for i in range(1, tries + 1):
        for target in alt_urls:
            try:
                log(f"[goto] tentativa {i}/{tries} -> {target}")
                page.goto(target, wait_until="load", timeout=60_000)
                return
            except Exception as e:
                log(f"[goto] falhou em {target}: {e}")
        page.wait_for_timeout(i * 3000)  # backoff progressivo (menor)
    page.goto(url, wait_until="load", timeout=60_000)

# ================ MAIN =================
def main():
    # Liga o watchdog ANTES de qualquer coisa pesada.
    start_watchdog(MAX_RUNTIME_SECONDS)

    ok = False  # <<< NOVO: controla se o fluxo terminou com sucesso de verdade

    with sync_playwright() as pw:
        context = pw.chromium.launch_persistent_context(
            user_data_dir=str(Path.cwd() / "pw_state"),
            headless=True,
            accept_downloads=True,
            ignore_https_errors=True,
            viewport={"width": 1600, "height": 950},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124 Safari/537.36",
        )
        page = context.pages[0] if context.pages else context.new_page()
        # timeouts reduzidos: 180s -> 30s (ação) e 300s -> 60s (navegação)
        page.set_default_timeout(30_000)
        page.set_default_navigation_timeout(60_000)

        page.on("console", lambda m: log(f"[console] {m.type}: {m.text}"))
        page.on("pageerror", lambda e: log(f"[pageerror] {e}"))
        page.on("requestfailed", lambda r: log(f"[requestfailed] {r.url} -> {r.failure}"))

        hook_filter_capture(page)
        hook_export_capture(page)

        try:
            step("1) Login")
            goto_login_with_retries(page, LOGIN_URL)

            if "/#/login" in page.url:
                page.fill(SEL["rede"], REDE)
                page.fill(SEL["email"], USER)
                page.fill(SEL["senha"], PASS)
                page.click(SEL["entrar"])

            page.wait_for_url("**/#/loja/**", timeout=120_000)
            try:
                page.wait_for_load_state("networkidle", timeout=15_000)
            except Exception:
                pass

            # 2) Trocar loja via secret (habilite se quiser forçar)
            trocar_loja(page, os.getenv("APPNEXT_LOJA_DESTINO"))

            # 3) Ir para Vendas > Vendedor Analítico
            step("3) Menu: Vendas → Vendedor Analítico")
            goto_vendedor_analitico_via_menu(page)
            wait_for_titulo_rel(page, timeout=40000)
            log("Tela confirmada.")

            # 4) Abrir filtros e estabilizar grid
            step("4) Abrindo filtros")
            open_filters_pane(page)
            try:
                page.locator(SEL["btn_atualizar"]).first.click(timeout=3000)
            except Exception:
                pass
            page.wait_for_selector(SEL["grid_row"], timeout=30000)

            # 5) Export
            step("5) Exportar Excel (tentar capturar URL)")
            captured_export_ok = False
            try:
                # timeout reduzido de 180s -> 90s
                with page.expect_download(timeout=90_000) as dl_tmp:
                    page.locator(SEL["excel_btn"]).first.click()
                tmp_download = dl_tmp.value
                tmp_path = Path.cwd() / f"_tmp_export_{TS}.xlsx"
                tmp_download.save_as(tmp_path.as_posix())
                log("Export padrão baixado (pode ser blob:/POST). Tentando reaproveitar URL capturada...")

                if CapturedExport.url:
                    step(f"5b) Reemitindo export com período {FMT_BR_INI} → {FMT_BR_FIM}")
                    replay_export_with_dates(page, LOCAL_OUT)
                    captured_export_ok = True
                else:
                    log("Não consegui capturar URL direta do Excel (provável blob:). Usarei fallback via API.")

                try:
                    tmp_path.unlink(missing_ok=True)
                except Exception:
                    pass
            except Exception as e:
                log(f"Aviso: falha ao acionar o botão Excel ou capturar URL: {e}. Vou usar fallback via API.")

            if not captured_export_ok:
                step(f"5c) Fallback: baixando dados da API {FMT_BR_INI} → {FMT_BR_FIM} e gerando Excel")
                data = fetch_report_json_with_dates(page)
                write_excel_from_json(data, LOCAL_OUT)

            # 6) Upload no Drive
            step("6) Enviando para Google Drive")
            rclone_copy_latest(LOCAL_OUT)
            log("Upload concluído.")

            ok = True  # <<< só chega aqui se TUDO acima deu certo

        except Exception:
            log("### ERRO DURANTE O FLUXO ###")
            print(traceback.format_exc())
        finally:
            try:
                context.close()
            except Exception:
                pass

    step("FINALIZADO")
    log(f"Local: {LOCAL_OUT.resolve()}")
    log(f"Drive: {DRIVE_FILE_NAME}  (pasta ID {DRIVE_FOLDER_ID})")

    # <<< NOVO: faz o job FALHAR de verdade se algo deu errado.
    # Antes, qualquer erro era "engolido" e o Actions mostrava verde.
    if not ok:
        log("### Finalizando com erro (exit 1) ###")
        sys.exit(1)


if __name__ == "__main__":
    main()
